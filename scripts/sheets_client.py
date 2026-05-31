"""
Google Sheets API wrapper.
Handles all reads/writes to the trading data spreadsheet.
"""

import os
import json
import gspread
from google.oauth2.service_account import Credentials
from datetime import datetime, date
import pandas as pd

SCOPES = [
    "https://spreadsheets.google.com/feeds",
    "https://www.googleapis.com/auth/drive",
]

# ── Connection ────────────────────────────────────────────────────────────────

def get_client():
    """Authenticate and return a gspread client."""
    creds_json = os.environ.get("GOOGLE_CREDENTIALS")
    if not creds_json:
        raise RuntimeError("GOOGLE_CREDENTIALS environment variable not set")
    creds_dict = json.loads(creds_json)
    creds = Credentials.from_service_account_info(creds_dict, scopes=SCOPES)
    return gspread.authorize(creds)


def get_sheet(tab_name: str):
    """Open a specific tab in the trading spreadsheet."""
    client = get_client()
    spreadsheet_id = os.environ.get("SPREADSHEET_ID")
    if not spreadsheet_id:
        raise RuntimeError("SPREADSHEET_ID environment variable not set")
    wb = client.open_by_key(spreadsheet_id)
    try:
        return wb.worksheet(tab_name)
    except gspread.WorksheetNotFound:
        # Auto-create missing tabs
        return wb.add_worksheet(title=tab_name, rows=5000, cols=50)


# ── TOP GAINERS DATA ──────────────────────────────────────────────────────────

TOP_GAINERS_HEADERS = [
    "DATE", "DOW", "WEEK#", "STOCK", "FLOAT", "TOD",
    "TIME IN", "TIME OUT", "TRADE TIME", "ENTRY TYPE",
    "ENTRY PRICE", "EXIT PRICE",
    "20 MA", "200 MA", "STATE", "RANGE", "POSITION",
    "GAIN $/SHARE", "G/L %/SHARE", "# LEGS", "A+ OPP?",
    "MDR WATCHLIST", "NEWS Y/N", "NEWS CATEGORY", "NOTES",
]

def read_top_gainers(days_back: int = 90) -> pd.DataFrame:
    """Read TOP Gainers Data sheet, stripping spaces from column names.
    Uses get_all_records() so percentage-formatted cells return floats (0.49)
    not display strings ("49%"). Keys are then stripped of leading/trailing spaces.
    """
    from config import SHEET_TOP_GAINERS
    ws = get_sheet(SHEET_TOP_GAINERS)

    try:
        # get_all_records returns proper Python types (floats for numeric/pct cells)
        data = ws.get_all_records()
    except Exception:
        # Fallback to raw values if records call fails
        all_values = ws.get_all_values()
        if not all_values or len(all_values) < 2:
            return pd.DataFrame(columns=TOP_GAINERS_HEADERS)
        headers = [h.strip() for h in all_values[0]]
        data = [dict(zip(headers, row)) for row in all_values[1:]]

    if not data:
        return pd.DataFrame(columns=TOP_GAINERS_HEADERS)

    # Strip spaces from all column names: ' ENTRY PRICE ' → 'ENTRY PRICE'
    data = [{k.strip(): v for k, v in row.items()} for row in data]
    # Alias personal Excel column names → internal names used by scorer/generator
    _ALIAS = {"DOW":"DAY OF WEEK","TRADE TIME":"RUN TIME","STATE":"TYPE OF STATE","G/L %/SHARE":"GAIN %/SHARE"}
    data = [{_ALIAS.get(k,k): v for k, v in row.items()} for row in data]

    df = pd.DataFrame(data)
    if "DATE" in df.columns:
        df["DATE"] = pd.to_datetime(df["DATE"], errors="coerce")
        cutoff = pd.Timestamp.today() - pd.Timedelta(days=days_back)
        df = df[df["DATE"] >= cutoff]
    return df


def append_top_gainers_rows(rows: list[dict]):
    """
    Append new rows using SHEET HEADER ORDER (not positional).
    This handles any sheet column structure (26 or 30 cols) correctly.
    """
    from config import SHEET_TOP_GAINERS
    ws = get_sheet(SHEET_TOP_GAINERS)

    # Get ACTUAL sheet headers — use these as the write order
    sheet_headers = [h.strip() for h in ws.row_values(1)]

    # If sheet is empty, write our headers first
    if not sheet_headers:
        ws.append_row(TOP_GAINERS_HEADERS)
        sheet_headers = TOP_GAINERS_HEADERS[:]

    # Build existing keys for dedup
    existing = ws.get_all_records()
    existing_keys = set()
    for r in existing:
        # Normalize DATE to YYYY-MM-DD string (gspread may return datetime objects)
        d = r.get("DATE", "")
        if hasattr(d, 'strftime'):
            d = d.strftime("%Y-%m-%d")
        else:
            d = str(d).strip()[:10]
        key = (d, str(r.get("STOCK", "")).strip().upper(), str(r.get("TIME IN", "")).strip())
        existing_keys.add(key)

    new_rows = []
    added = 0
    for row in rows:
        key = (str(row.get("DATE", "")).strip()[:10], str(row.get("STOCK", "")).strip().upper(), str(row.get("TIME IN", "")).strip())
        if key in existing_keys:
            continue
        # Write in SHEET column order — unknown/extra columns get empty string
        ordered = [row.get(h, "") for h in sheet_headers]
        new_rows.append(ordered)
        existing_keys.add(key)
        added += 1

    if new_rows:
        ws.append_rows(new_rows, value_input_option="USER_ENTERED")

    print(f"  Appended {added} new rows to {SHEET_TOP_GAINERS} (sheet cols: {len(sheet_headers)})")
    return added


# ── MDR TRACKING ──────────────────────────────────────────────────────────────

MDR_HEADERS = [
    "STOCK", "INITIAL BO DATE", "MDR LIST DATE",
    "ENTRY TYPE", "ENTRY TIME", "EXIT TIME", "TRADE TIME",
    "FLOAT",
    "ENTRY PRICE", "EXIT PRICE",
    "# LEGS", "20 MA", "200 MA",
    "STATE", "RANGE", "POSITION",
    "GAIN $/SHARE", "GAIN %/SHARE",
    "MDR SCORE",
    "TIER", "DAYS ON LIST", "LAST RUN DATE",
    "NEWS TYPE", "NOTES",
]

def read_mdr_tracking() -> pd.DataFrame:
    """Read MDR TRACKING sheet."""
    from config import SHEET_MDR_TRACKING
    ws = get_sheet(SHEET_MDR_TRACKING)
    data = ws.get_all_records()
    if not data:
        return pd.DataFrame(columns=MDR_HEADERS)
    return pd.DataFrame(data)


def write_mdr_tracking(df: pd.DataFrame, removed_tickers: set = None):
    """
    Upsert MDR TRACKING: update scored stocks, preserve unscored ones,
    remove only explicitly excluded stocks. Never wipes the full sheet.
    """
    from config import SHEET_MDR_TRACKING
    ws = get_sheet(SHEET_MDR_TRACKING)
    if removed_tickers is None:
        removed_tickers = set()

    # Read existing full sheet
    try:
        existing_values = ws.get_all_values()
    except Exception:
        existing_values = []

    # Empty sheet — write fresh
    if not existing_values or not any(existing_values[0]):
        all_rows = [MDR_HEADERS]
        for _, row in df.iterrows():
            ordered = [str(row.get(h, "") or "") for h in MDR_HEADERS]
            all_rows.append(ordered)
        ws.update(all_rows, value_input_option="USER_ENTERED")
        print(f"  MDR TRACKING written fresh: {len(df)} rows")
        return

    # Parse existing sheet
    existing_headers = [h.strip() for h in existing_values[0]]
    stock_col = next((i for i, h in enumerate(existing_headers) if h.upper() == "STOCK"), 0)

    # Map existing rows by ticker
    existing_by_ticker = {}
    for row in existing_values[1:]:
        if not row or len(row) <= stock_col:
            continue
        ticker = str(row[stock_col]).strip().upper()
        if not ticker:
            continue
        padded = list(row) + [""] * max(0, len(existing_headers) - len(row))
        existing_by_ticker[ticker] = dict(zip(existing_headers, padded))

    # Map updated/scored rows by ticker
    updated_by_ticker = {}
    for _, row in df.iterrows():
        ticker = str(row.get("STOCK", "") or "").strip().upper()
        if ticker:
            updated_by_ticker[ticker] = row.to_dict()

    # Build final row list BEFORE touching the sheet (crash safety)
    final_rows = [MDR_HEADERS]
    seen = set()

    # Add scored/updated stocks first
    for _, row in df.iterrows():
        ticker = str(row.get("STOCK", "") or "").strip().upper()
        if ticker and ticker not in removed_tickers:
            ordered = [("" if (row.get(h) is None or str(row.get(h,"")).strip() == "" or str(row.get(h,"")).strip() == "nan") else str(row.get(h,""))) for h in MDR_HEADERS]
            final_rows.append(ordered)
            seen.add(ticker)

    # Preserve existing stocks that were not scored and not removed
    for ticker, row_dict in existing_by_ticker.items():
        if ticker not in seen and ticker not in removed_tickers:
            ordered = [("" if (row_dict.get(h) is None or str(row_dict.get(h,"")).strip() == "" or str(row_dict.get(h,"")).strip() == "nan") else str(row_dict.get(h,""))) for h in MDR_HEADERS]
            final_rows.append(ordered)
            seen.add(ticker)

    # Clear and write atomically — clear only after successfully building rows
    try:
        ws.clear()
        ws.update(final_rows, value_input_option="USER_ENTERED")
        preserved = len(seen) - len(updated_by_ticker)
        print(f"  MDR TRACKING: {len(updated_by_ticker)} scored + {preserved} preserved = {len(final_rows)-1} total ({len(removed_tickers)} removed)")
    except Exception as e:
        print(f"  MDR TRACKING write error: {e} — sheet may need manual restore")
        raise


def upsert_mdr_stock(stock_data: dict):
    """Add or update a stock in the MDR TRACKING sheet."""
    df = read_mdr_tracking()
    ticker = stock_data.get("STOCK", "")
    if ticker in df["STOCK"].values:
        idx = df.index[df["STOCK"] == ticker][0]
        for k, v in stock_data.items():
            if k in df.columns:
                df.at[idx, k] = v
    else:
        new_row = {h: stock_data.get(h, "") for h in MDR_HEADERS}
        df = pd.concat([df, pd.DataFrame([new_row])], ignore_index=True)
    write_mdr_tracking(df)


# ── SCAN LOG ──────────────────────────────────────────────────────────────────

SCAN_LOG_HEADERS = ["DATE", "TICKER", "SOURCE", "TIMESTAMP", "GAIN_PCT", "PRICE"]

def log_tickers(tickers: list[dict], source: str):
    """Write ticker discoveries to SCAN LOG."""
    from config import SHEET_SCAN_LOG
    ws = get_sheet(SHEET_SCAN_LOG)
    existing = ws.get_all_records()
    if not existing:
        ws.append_row(SCAN_LOG_HEADERS)

    today = date.today().isoformat()
    ts = datetime.now().strftime("%H:%M")
    for t in tickers:
        row = [
            today,
            t.get("ticker", ""),
            source,
            ts,
            t.get("gain_pct", ""),
            t.get("price", ""),
        ]
        ws.append_row(row, value_input_option="USER_ENTERED")


def read_today_scan_log() -> list[str]:
    """Return unique tickers logged today. Handles missing headers and date formats."""
    from config import SHEET_SCAN_LOG
    ws = get_sheet(SHEET_SCAN_LOG)

    # Auto-create headers if sheet is empty
    all_vals = ws.get_all_values()
    if not all_vals or not any(all_vals[0]):
        ws.update([["DATE", "TICKER", "SOURCE", "TIMESTAMP", "GAIN_PCT", "PRICE"]])
        return []

    # Read records
    try:
        data = ws.get_all_records()
    except Exception:
        return []
    if not data:
        return []

    # Match against both today AND the most recent trading day
    # so tickers added yesterday still get picked up on midnight runs
    import pytz
    from datetime import timedelta
    _et  = pytz.timezone("America/New_York")
    _now = __import__("datetime").datetime.now(_et)
    _d   = _now.date()
    if _now.hour < 16:
        _d -= timedelta(days=1)
    while _d.weekday() >= 5:
        _d -= timedelta(days=1)
    valid_dates = {date.today().isoformat(), _d.isoformat()}

    tickers = []
    seen = set()
    for r in data:
        raw_date = str(r.get("DATE", "") or "").strip()[:10]
        ticker   = str(r.get("TICKER", "") or "").strip().upper()
        if raw_date in valid_dates and ticker and ticker not in seen:
            tickers.append(ticker)
            seen.add(ticker)
    return tickers


# ── EXPORT ────────────────────────────────────────────────────────────────────

def export_all_to_excel(output_path: str):
    """Export all sheets to a single Excel file for download."""
    from config import SHEET_TOP_GAINERS, SHEET_MDR_TRACKING, SHEET_SCAN_LOG
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default sheet

    BG_HDR = "0D1117"
    TXT_GLD = "E3B341"

    for tab_name in [SHEET_TOP_GAINERS, SHEET_MDR_TRACKING, SHEET_SCAN_LOG]:
        try:
            ws_src = get_sheet(tab_name)
            data = ws_src.get_all_values()
            ws = wb.create_sheet(title=tab_name)
            for r_idx, row in enumerate(data, 1):
                for c_idx, val in enumerate(row, 1):
                    cell = ws.cell(row=r_idx, column=c_idx, value=val)
                    if r_idx == 1:
                        cell.font = Font(color=TXT_GLD, bold=True, name="Arial", size=10)
                        cell.fill = PatternFill("solid", fgColor=BG_HDR)
                        cell.alignment = Alignment(horizontal="center")
        except Exception as e:
            print(f"  Warning: could not export {tab_name}: {e}")

    wb.save(output_path)
    print(f"  Exported all data → {output_path}")
